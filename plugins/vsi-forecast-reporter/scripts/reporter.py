#!/usr/bin/env python3
"""
VSI Forecast Reporter — 核心引擎

流程:
  Step 0: 解析所有 forecast 檔，攤平成 long table
  Step 1: 轉成 wide table (M / M+1 / M+2 三個月一列)
  Step 2: 用檔名辨識三份 Allocate Table 的月份 (M / M+1 / M+2)
  Step 3: 找 Missing Model (forecast 有 Qty 但 Allocate 沒此列)
  Step 4: 找 Missing ASP (有 Qty 但 ASP 為 0/空)
  Step 5: 產出三份輸出檔
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


# ============================================================
# 設定載入
# ============================================================

def load_configs(config_dir: Path):
    """載入所有 product_line config 與通用 rules"""
    pl_configs = []
    for f in sorted((config_dir / "product_lines").glob("*.json")):
        if f.name.startswith("_"):
            continue
        with open(f, encoding="utf-8") as fp:
            pl_configs.append(json.load(fp))
    with open(config_dir / "rules.json", encoding="utf-8") as fp:
        rules = json.load(fp)
    return pl_configs, rules


# ============================================================
# 產品線辨識
# ============================================================

def detect_product_line(filepath: Path, configs: list[dict]) -> dict | None:
    """先用檔名 regex 辨識，找不到再用 sheet 結構特徵"""
    name = filepath.name

    # 1) 檔名匹配
    for cfg in configs:
        for pattern in cfg.get("filename_patterns", []):
            if re.search(pattern, name, re.IGNORECASE):
                return cfg

    # 2) 內容匹配
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for cfg in configs:
            sig = cfg.get("content_signature")
            if not sig:
                continue
            country_sheets = cfg["country_sheets"]
            sample = next((s for s in country_sheets if s in wb.sheetnames), None)
            if not sample:
                continue
            ws = wb[sample]
            row_text = " | ".join(
                str(ws.cell(row=sig["row"], column=c).value or "")
                for c in range(1, 16)
            ).upper()
            if all(kw.upper() in row_text for kw in sig["must_contain"]):
                return cfg
    except Exception as e:
        print(f"  (內容辨識失敗 {filepath.name}: {e})", file=sys.stderr)

    return None


# ============================================================
# Forecast 解析
# ============================================================

def parse_country_sheet(ws, layout, dim_cols, metric_offsets, country, product_line, num_months, local_currency_countries):
    """從一個國家分頁解析出 long format records

    依 country 決定 gross_asp 取自 USD 或 Local 欄
    """
    rows = []
    data_start = layout["data_start_row"]
    cols_per_month = layout["cols_per_month"]
    gap = layout["gap_cols_between_blocks"]
    first_col = layout["first_block_start_col"]
    block_stride = cols_per_month + gap
    model_col = dim_cols["model"]

    # 決定這個國家用哪個 ASP 欄
    use_local = country in local_currency_countries
    asp_offset_key = "gross_asp_local" if use_local else "gross_asp_usd"

    for row_idx in range(data_start, ws.max_row + 1):
        # 結束標記：A/B/C 欄出現 Total / Subtotal 就停
        stop = False
        for c in (1, 2, 3):
            v = ws.cell(row=row_idx, column=c).value
            if v and str(v).strip().lower() in ("total", "subtotal", "grand total"):
                stop = True
                break
        if stop:
            break

        model = ws.cell(row=row_idx, column=model_col).value
        if not model or not str(model).strip():
            continue
        model = str(model).strip()

        # 過濾分類標籤 (純文字短字串、無數字/連字號)
        if not any(ch.isdigit() or ch in "-_+" for ch in model) and len(model) < 8:
            continue

        # 抓三個月的 Qty / Gross ASP
        record = {
            "product_line": product_line,
            "country": country,
            "model": model,
        }
        has_any_data = False
        for m in range(num_months):
            block_start = first_col + m * block_stride
            qty = ws.cell(row=row_idx, column=block_start + metric_offsets["qty"]).value
            asp = ws.cell(row=row_idx, column=block_start + metric_offsets[asp_offset_key]).value

            month_label = "M" if m == 0 else f"M+{m}"
            record[f"{month_label}_qty"] = _round2(qty)
            record[f"{month_label}_gross_asp"] = _round2(asp)

            if qty or asp:
                has_any_data = True

        if has_any_data:
            rows.append(record)

    return rows


def parse_forecast_file(filepath: Path, config: dict, num_months: int, local_currency_countries: list) -> pd.DataFrame:
    """解析整份 forecast 檔，回傳 wide-format DataFrame"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    product_line = config["product_line"]
    layout = config["country_layout"]
    dim_cols = config["dimension_columns"]
    metric_offsets = config["metric_offsets"]
    country_sheets = config["country_sheets"]

    all_rows = []
    for country in country_sheets:
        if country not in wb.sheetnames:
            print(f"  ⚠ [{product_line}] 找不到分頁: {country}（跳過）", file=sys.stderr)
            continue
        ws = wb[country]
        rows = parse_country_sheet(
            ws, layout, dim_cols, metric_offsets,
            country, product_line, num_months, local_currency_countries
        )
        all_rows.extend(rows)
        print(f"  ✓ [{product_line}] {country}: {len(rows)} 個機種", file=sys.stderr)

    return pd.DataFrame(all_rows)


# ============================================================
# Allocate Table 處理
# ============================================================

def detect_allocate_month(filepath: Path, pattern: str) -> tuple[int, int] | None:
    """從檔名辨識 (year, month)"""
    m = re.search(pattern, filepath.name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def load_allocate_table(filepath: Path, rules: dict) -> pd.DataFrame:
    """讀取 Allocate Table，回傳含 country/model/currency 的 DataFrame"""
    alloc_cfg = rules["allocate_table"]
    df = pd.read_excel(filepath, sheet_name=alloc_cfg["sheet_name"])
    df.columns = [str(c).strip() for c in df.columns]
    # 保留必要欄位
    df = df.rename(columns={
        alloc_cfg["key_columns"]["country"]: "country",
        alloc_cfg["key_columns"]["model"]: "model",
    })
    df["country"] = df["country"].astype(str).str.strip()
    df["model"] = df["model"].astype(str).str.strip()
    return df


def sort_allocate_files(allocate_files: list[Path], rules: dict) -> list[tuple[Path, int, int]]:
    """依檔名抓出年月並排序 → [(file, year, month), ...]"""
    pattern = rules["allocate_table"]["filename_pattern"]
    out = []
    for f in allocate_files:
        ym = detect_allocate_month(f, pattern)
        if ym is None:
            print(f"  ⚠ 無法從檔名辨識年月: {f.name}（跳過）", file=sys.stderr)
            continue
        out.append((f, ym[0], ym[1]))
    out.sort(key=lambda x: (x[1], x[2]))
    return out


# ============================================================
# Missing 偵測
# ============================================================

def find_missing(master_df: pd.DataFrame, allocate_files_sorted: list, rules: dict):
    """
    回傳:
      missing_models: forecast 有 Qty>0 但 Allocate 沒有此 (country, model) 的列
      missing_asp:    任何月份 Qty>0 但 Gross ASP=0/null 的列
    """
    skip_countries = set(rules.get("skip_countries", []))
    num_months = rules["output_months"]

    # 月份標籤 (M, M+1, M+2)
    month_labels = ["M"] + [f"M+{i}" for i in range(1, num_months)]

    # 讀取所有 Allocate Table 並建立 (country, model) 集合
    allocate_keys_by_month = {}  # month_index → set of (country, model)

    for idx, (filepath, year, month) in enumerate(allocate_files_sorted):
        if idx >= num_months:
            break
        df = load_allocate_table(filepath, rules)
        keys = set()
        for _, r in df.iterrows():
            country = r["country"]
            model = r["model"]
            if country in skip_countries:
                continue
            keys.add((country, model))
        allocate_keys_by_month[idx] = keys
        print(f"  ✓ Allocate {year}-{month:02d} ({filepath.name}): {len(keys)} 個 (Country, Model)", file=sys.stderr)

    # ============= Missing Models =============
    missing_models_records = []
    for _, row in master_df.iterrows():
        country = row["country"]
        model = row["model"]
        if country in skip_countries:
            continue

        for m_idx, label in enumerate(month_labels):
            qty = row.get(f"{label}_qty")
            if not _has_qty(qty):
                continue
            if m_idx not in allocate_keys_by_month:
                continue
            if (country, model) not in allocate_keys_by_month[m_idx]:
                missing_models_records.append({
                    "month": label,
                    "product_line": row["product_line"],
                    "country": country,
                    "model": model,
                    "qty": qty,
                    "gross_asp": row.get(f"{label}_gross_asp"),
                    "reason": "Not in Allocate Table",
                })

    # ============= Missing ASP =============
    missing_asp_records = []
    for _, row in master_df.iterrows():
        country = row["country"]
        model = row["model"]
        if country in skip_countries:
            continue

        for m_idx, label in enumerate(month_labels):
            qty = row.get(f"{label}_qty")
            asp = row.get(f"{label}_gross_asp")
            if not _has_qty(qty):
                continue
            if not _has_value(asp):
                missing_asp_records.append({
                    "month": label,
                    "product_line": row["product_line"],
                    "country": country,
                    "model": model,
                    "qty": qty,
                    "gross_asp": asp,
                    "reason": "Qty > 0 but Gross ASP missing",
                })

    return pd.DataFrame(missing_models_records), pd.DataFrame(missing_asp_records)


# ============================================================
# 回填 Allocate Tables
# ============================================================

def fill_allocate_tables(master_df: pd.DataFrame, allocate_files_sorted: list,
                         rules: dict, output_dir: Path) -> list[Path]:
    """
    把 master_df 的 qty / gross_asp 回填到三份 Allocate Table

    規則:
    - (Country, Model) 為比對鍵
    - Feb 用 M, Mar 用 M+1, Apr 用 M+2 (依排序)
    - 完全以 Master 為準覆寫 (Allocate 原值會被洗掉)
    - Master 有 Qty 但無 Gross ASP → Model Price 填 0
    - Country 在 skip_countries (TW) → 跳過該列
    - Master 沒提到的 (Country, Model) → 不動 Allocate 那列 (保留原值)

    回傳: 產出的檔案路徑清單
    """
    num_months = rules["output_months"]
    skip_countries = set(rules.get("skip_countries", []))
    alloc_cfg = rules["allocate_table"]

    # 建 Master 索引: (country, model) → row
    master_index = {}
    for _, r in master_df.iterrows():
        master_index[(r["country"], r["model"])] = r

    out_paths = []
    month_labels = ["M"] + [f"M+{i}" for i in range(1, num_months)]

    for idx, (src_path, year, month) in enumerate(allocate_files_sorted):
        if idx >= num_months:
            break
        label = month_labels[idx]
        qty_key = f"{label}_qty"
        asp_key = f"{label}_gross_asp"

        # 複製原檔到輸出目錄再修改 (不動原檔)
        out_path = output_dir / f"{src_path.stem}_filled.xlsx"
        wb = openpyxl.load_workbook(src_path)
        ws = wb[alloc_cfg["sheet_name"]]

        # 找欄位編號
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=alloc_cfg["header_row"], column=col).value
            if h:
                headers[str(h).strip()] = col

        country_col = headers["Country"]
        model_col = headers["Model"]
        price_col = headers["Model Price"]
        qty_col = headers["Forecase Qty"]

        filled_count = 0
        skipped_country_count = 0
        not_in_master_count = 0

        for row_idx in range(alloc_cfg["data_start_row"], ws.max_row + 1):
            country = ws.cell(row=row_idx, column=country_col).value
            model = ws.cell(row=row_idx, column=model_col).value
            if not country or not model:
                continue
            country = str(country).strip()
            model = str(model).strip()

            if country in skip_countries:
                skipped_country_count += 1
                continue

            master_row = master_index.get((country, model))
            if master_row is None:
                # Master 沒提到 → 不動
                not_in_master_count += 1
                continue

            # 以 Master 為準覆寫
            qty_val = master_row.get(qty_key)
            asp_val = master_row.get(asp_key)

            # Qty 處理 (整數)
            if pd.notna(qty_val) and qty_val is not None:
                try:
                    qty_int = int(round(float(qty_val)))
                    ws.cell(row=row_idx, column=qty_col).value = qty_int
                except (TypeError, ValueError):
                    pass

            # Price 處理: Master 有 Qty 但無 ASP → 填 0
            if _has_qty(qty_val) and not _has_value(asp_val):
                ws.cell(row=row_idx, column=price_col).value = 0
            elif pd.notna(asp_val) and asp_val is not None:
                try:
                    price_rounded = round(float(asp_val), 2)
                    ws.cell(row=row_idx, column=price_col).value = price_rounded
                except (TypeError, ValueError):
                    pass

            filled_count += 1

        wb.save(out_path)
        out_paths.append(out_path)
        print(f"  ✓ {label} ({year}-{month:02d}) → {out_path.name}:", file=sys.stderr)
        print(f"      已填 {filled_count} 列；TW 跳過 {skipped_country_count} 列；Master 未提到 {not_in_master_count} 列", file=sys.stderr)

    return out_paths


def _has_qty(v):
    """有意義的 Qty: 非 None / 非 0"""
    if v is None:
        return False
    try:
        return float(v) > 0
    except (TypeError, ValueError):
        return False


def _has_value(v):
    """有意義的數值 (price 用): 非 None / 非 0 / 非空字串"""
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    try:
        return float(v) > 0
    except (TypeError, ValueError):
        return False


def _round2(v):
    """四捨五入到小數第 2 位。None / 字串 / 0 維持原樣。"""
    if v is None:
        return None
    if isinstance(v, str):
        return v
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return v


# ============================================================
# 輸出格式化
# ============================================================

def write_master_excel(df: pd.DataFrame, path: Path, num_months: int):
    """產出 master_forecast.xlsx, 標題分組美化"""
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Master", index=False)

    # 美化
    wb = openpyxl.load_workbook(path)
    ws = wb["Master"]

    # 標題列粗體 + 底色
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 月份分組用顏色標 (M=綠 / M+1=黃 / M+2=橘)
    month_colors = ["E2EFDA", "FFF2CC", "FCE4D6"]
    headers = [c.value for c in ws[1]]
    for col_idx, h in enumerate(headers, 1):
        if not h:
            continue
        for m in range(num_months):
            prefix = "M" if m == 0 else f"M+{m}"
            if str(h).startswith(prefix + "_"):
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).fill = PatternFill("solid", start_color=month_colors[m])
                break

    # 自動寬度
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 25)

    ws.freeze_panes = "D2"  # 凍結 product_line / country / model 三欄
    wb.save(path)


def write_questionnaire_excel(df: pd.DataFrame, path: Path, title: str, instructions: str):
    """產出給 BLM/CSC 填寫的問卷檔，標頭加說明"""
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name=title, index=False, startrow=2)

    wb = openpyxl.load_workbook(path)
    ws = wb[title]

    # 加說明列
    ws.cell(row=1, column=1).value = f"📋 {title}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="305496")
    ws.cell(row=2, column=1).value = instructions
    ws.cell(row=2, column=1).font = Font(italic=True, color="595959")

    # 標題列美化
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 自動寬度
    for col in ws.columns:
        vals = [c.value for c in col if c.value is not None]
        if vals:
            max_len = max(len(str(v)) for v in vals)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 30)

    ws.freeze_panes = "A4"
    wb.save(path)


# ============================================================
# Main
# ============================================================

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--forecasts", nargs="+", required=True, help="廠商 forecast 檔（一個或多個）")
    p.add_argument("--allocates", nargs="+", required=True, help="Allocate Table 檔（M / M+1 / M+2，可亂序）")
    p.add_argument("--config-dir", required=True, help="config 目錄")
    p.add_argument("--output-dir", default="./output", help="輸出目錄")
    args = p.parse_args()

    config_dir = Path(args.config_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    pl_configs, rules = load_configs(config_dir)
    num_months = rules["output_months"]
    local_currency_countries = rules.get("local_currency_countries", [])
    print(f"[1/5] 載入設定", file=sys.stderr)
    print(f"  產品線: {[c['product_line'] for c in pl_configs]}", file=sys.stderr)
    print(f"  輸出月份: M, M+1, ..., M+{num_months-1}", file=sys.stderr)
    print(f"  用 Local 幣別的國家: {local_currency_countries} (其他國家用 USD)", file=sys.stderr)

    print(f"\n[2/5] 解析 forecast 檔案", file=sys.stderr)
    all_dfs = []
    for fpath in args.forecasts:
        fpath = Path(fpath)
        cfg = detect_product_line(fpath, pl_configs)
        if cfg is None:
            print(f"  ✗ 無法辨識產品線: {fpath.name}", file=sys.stderr)
            continue
        print(f"  → {fpath.name} = {cfg['product_line']}", file=sys.stderr)
        df = parse_forecast_file(fpath, cfg, num_months, local_currency_countries)
        all_dfs.append(df)

    if not all_dfs:
        print("沒有可解析的 forecast 檔，中止", file=sys.stderr)
        sys.exit(1)

    master_df = pd.concat(all_dfs, ignore_index=True)
    master_df = master_df.sort_values(["product_line", "country", "model"]).reset_index(drop=True)
    print(f"\n  合計機種列數: {len(master_df)}", file=sys.stderr)

    print(f"\n[3/5] 比對 Allocate Tables", file=sys.stderr)
    allocate_paths = [Path(p) for p in args.allocates]
    allocate_sorted = sort_allocate_files(allocate_paths, rules)
    print(f"  辨識結果 (依年月排序):", file=sys.stderr)
    for idx, (f, y, m) in enumerate(allocate_sorted):
        label = "M" if idx == 0 else f"M+{idx}"
        print(f"    {label} = {y}-{m:02d}  ({f.name})", file=sys.stderr)

    missing_models_df, missing_asp_df = find_missing(master_df, allocate_sorted, rules)

    print(f"\n[4/5] 輸出檔案", file=sys.stderr)
    out_names = rules["output_filenames"]
    master_path = output_dir / out_names["master"]
    missing_models_path = output_dir / out_names["missing_models"]
    missing_asp_path = output_dir / out_names["missing_asp"]

    write_master_excel(master_df, master_path, num_months)
    print(f"  ✓ Master: {master_path.name} ({len(master_df)} 列)", file=sys.stderr)

    write_questionnaire_excel(
        missing_models_df, missing_models_path, "Missing_Models",
        "以下機種有 Qty 但不在對應月份的 Allocate Table 中。請 BLM 審查後決定是否加入 Allocate。"
    )
    print(f"  ✓ Missing Models: {missing_models_path.name} ({len(missing_models_df)} 列)", file=sys.stderr)

    write_questionnaire_excel(
        missing_asp_df, missing_asp_path, "Missing_ASP",
        "以下機種有 Qty 但對應的 Gross ASP 為空。請 CSC 向 BLM/NS 詢價後填入 gross_asp 欄位。"
    )
    print(f"  ✓ Missing ASP: {missing_asp_path.name} ({len(missing_asp_df)} 列)", file=sys.stderr)

    # 回填 Allocate Tables
    print(f"\n[5/5] 回填 Allocate Tables (以 Master 為準覆寫)", file=sys.stderr)
    filled_paths = fill_allocate_tables(master_df, allocate_sorted, rules, output_dir)

    print(f"\n✅ 完成！輸出在 {output_dir}", file=sys.stderr)
    print(f"  總共產出 {3 + len(filled_paths)} 份檔案:", file=sys.stderr)
    print(f"    1. {master_path.name}", file=sys.stderr)
    print(f"    2. {missing_models_path.name}", file=sys.stderr)
    print(f"    3. {missing_asp_path.name}", file=sys.stderr)
    for i, p in enumerate(filled_paths, 4):
        print(f"    {i}. {p.name}", file=sys.stderr)


if __name__ == "__main__":
    main()
